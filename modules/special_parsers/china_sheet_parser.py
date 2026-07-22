"""
中国 Sheet 特殊解析器

处理包含 "中国" sheet 的特殊 Excel 文件结构：
- 只解析 "English" 列的内容
- 读取公式计算后的值（而非公式本身）
- 翻译结果填入 "Landessprache / Local Language" 列
"""

from typing import List, Dict, Tuple, Optional


CHINA_SHEET_NAMES = ("中国", "中文", "CHINESE", "Chinese", "chinese", "中文翻译")


def find_china_sheet_name(sheet_names) -> Optional[str]:
    """Return the workbook sheet name that uses the China/local-language layout."""
    for candidate in CHINA_SHEET_NAMES:
        if candidate in sheet_names:
            return candidate
    return None


def parse_china_sheet(file_path: str, sheet_name: str = None) -> Tuple[List[Dict], Dict]:
    """
    解析特殊结构的 "中国" sheet
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        texts: 需要翻译的文本块列表
        format_info: 文件格式信息
        
    Raises:
        ValueError: 当找不到必要的列时
    """
    from openpyxl import load_workbook
    
    # 使用 data_only=True 打开文件，获取公式计算后的值
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"无法打开 Excel 文件: {e}")

    try:
        source_sheet = sheet_name or find_china_sheet_name(wb.sheetnames)
        if not source_sheet:
            raise ValueError('文件中没有支持的中文/中国 sheet')
        
        ws = wb[source_sheet]
        texts = []
        index = 0
        
        # 找到标题行，确定列位置
        english_col = None
        local_lang_col = None
        header_row = 0
        
        # 遍历前5行查找标题
        for row_idx in range(min(5, ws.max_row)):
            row = list(ws.iter_rows(min_row=row_idx+1, max_row=row_idx+1))[0]
            for col_idx, cell in enumerate(row):
                if cell.value:
                    header_text = str(cell.value).strip()
                    if header_text == "English":
                        english_col = col_idx
                    elif header_text == "Landessprache / Local Language":
                        local_lang_col = col_idx
            
            # 如果找到了两个列，停止搜索
            if english_col is not None and local_lang_col is not None:
                header_row = row_idx
                break
        
        # 如果找不到标题列，使用默认的 B 列和 C 列
        if english_col is None:
            english_col = 1  # B 列（索引从0开始，所以1是B列）
            print(f'[ChinaSheet] 未找到 "English" 列，使用默认 B 列 (列索引: {english_col})')

        if local_lang_col is None:
            local_lang_col = 2  # C 列（索引从0开始，所以2是C列）
            print(f'[ChinaSheet] 未找到 "Landessprache / Local Language" 列，使用默认 C 列 (列索引: {local_lang_col})')
        
        # 提取 "English" 列的数据
        for row_idx in range(header_row + 1, ws.max_row):
            english_cell = ws.cell(row=row_idx + 1, column=english_col + 1)
            
            # data_only=True 读取的是公式计算后的值
            if english_cell.value is not None:
                cell_text = str(english_cell.value).strip()
                # 跳过标题本身和空值
                if cell_text and cell_text != "English":
                    texts.append({
                        "text": cell_text,
                        "sheet": source_sheet,
                        "row": row_idx,
                        "col": local_lang_col,  # 目标列
                        "source_col": english_col,  # 源列
                        "index": index,
                        "is_formula": False,  # 已经是计算后的值
                        "format": {},  # 简化格式信息
                        "is_china_sheet": True  # 特殊结构标记
                    })
                    index += 1
        
        wb_info = {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "cell_count": len(texts),
            "formula_count": 0,
            "special_structure": "china_sheet",
            "source_sheet": source_sheet,
            "english_col": english_col,
            "local_lang_col": local_lang_col,
            "header_row": header_row
        }
        
        return texts, wb_info
    finally:
        wb.close()
