# -*- coding: utf-8 -*-
"""
文件解析模块测试
"""

import os
import sys
import unittest
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.file_parser import (
    parse_excel_file,
    parse_txt_file,
    parse_csv_file,
    get_file_type
)


class TestFileParser(unittest.TestCase):
    """测试文件解析功能"""
    
    def setUp(self):
        """测试前准备"""
        self.temp_dir = tempfile.mkdtemp()
    
    def tearDown(self):
        """测试后清理"""
        import gc
        import shutil
        gc.collect()
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def test_parse_txt_file(self):
        """测试TXT文件解析"""
        # 创建测试文件
        test_file = os.path.join(self.temp_dir, "test.txt")
        with open(test_file, "w", encoding="utf-8") as f:
            f.write("Hello World\n")
            f.write("Second Line\n")
        
        texts, info = parse_txt_file(test_file)
        
        self.assertEqual(len(texts), 2)
        self.assertEqual(texts[0]["text"], "Hello World")
        self.assertEqual(texts[1]["text"], "Second Line")
        self.assertEqual(info["encoding"], "utf-8")
    
    def test_parse_csv_file(self):
        """测试CSV文件解析"""
        import csv
        
        # 创建测试文件
        test_file = os.path.join(self.temp_dir, "test.csv")
        with open(test_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Hello", "World"])
            writer.writerow(["Second", "Row"])
        
        texts, info = parse_csv_file(test_file)
        
        self.assertEqual(len(texts), 4)
        self.assertEqual(texts[0]["text"], "Hello")
        self.assertEqual(texts[0]["row"], 0)
        self.assertEqual(texts[0]["col"], 0)

    def test_parse_excel_uses_special_parser_for_chinese_sheet(self):
        """测试中文sheet也使用中国表特殊解析逻辑"""
        from openpyxl import Workbook

        test_file = os.path.join(self.temp_dir, "chinese_sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "中文"
        ws.append(["ID", "English", "Landessprache / Local Language"])
        ws.append([1, "Problem solving sheet", ""])
        ws.append([2, "Root cause", ""])
        wb.save(test_file)
        wb.close()

        texts, info = parse_excel_file(test_file)

        self.assertEqual(info["special_structure"], "china_sheet")
        self.assertEqual(info["source_sheet"], "中文")
        self.assertEqual([item["text"] for item in texts], ["Problem solving sheet", "Root cause"])
        self.assertTrue(all(item["sheet"] == "中文" for item in texts))

    def test_parse_excel_skips_abnormally_large_empty_used_range(self):
        """测试异常大的空白Used Range不会触发全表扫描"""
        from openpyxl import Workbook

        test_file = os.path.join(self.temp_dir, "large_used_range.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Large"
        ws["A1"] = "Keep me"
        ws["B1"] = " "
        wb.save(test_file)
        wb.close()

        from unittest.mock import patch
        with patch("modules.file_parser.MAX_STANDARD_EXCEL_CELLS", 1):
            texts, info = parse_excel_file(test_file)

        self.assertEqual([item["text"] for item in texts], ["Keep me"])
        self.assertIn("Large", info.get("skipped_large_sheets", []))
    
    def test_get_file_type(self):
        """测试文件类型识别"""
        self.assertEqual(get_file_type("test.docx"), "word")
        self.assertEqual(get_file_type("test.xlsx"), "excel")
        self.assertEqual(get_file_type("test.pdf"), "pdf")
        self.assertEqual(get_file_type("test.txt"), "txt")
        self.assertEqual(get_file_type("test.csv"), "csv")
        self.assertEqual(get_file_type("test.unknown"), "unknown")


if __name__ == "__main__":
    unittest.main()
